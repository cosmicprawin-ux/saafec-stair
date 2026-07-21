#!/usr/bin/env python3
"""Mutation-table and cached SaProt embedding loader for SAAFEC-STAIR inference.

CSV and XLSX tables are normalized to per-protein mutation requests. Residue
identifiers are resolved against the cached PDB sequence before requested
substitutions are marked for prediction.
"""
from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, NamedTuple

import torch
from torch.utils.data import Dataset

from core.amino_acids import AA_TO_INDEX, AMINO_ACIDS_20, NUM_AMINO_ACIDS
from core.SAAFEC_STAIR_pdb_chain_parser import (
    build_structure_parser_context,
    infer_chain_from_stem,
    resolve_reference_mutation_index,
    safe_first_atom_chain,
    unique_chain_candidates,
)
from core.pipeline_config import (
    WORK_DIR,
    resolve_output_path,
    resolve_path_with_base,
)
from core.structure_identifiers import pdb_stem, structure_key

# ---------------------------------------------------------------------------
# Directory layout
# ---------------------------------------------------------------------------
INFERENCE_TABLE_BASE_DIR = WORK_DIR

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

class ProteinSample(NamedTuple):
    """Single whole-protein inference sample."""
    embedding: torch.Tensor    # (L, embedding_dim)
    ca_coordinates: torch.Tensor  # (L, 3), NaN rows when C-alpha coordinates are missing
    mutation_mask: torch.Tensor  # (L, 20)
    wt_sequence: str
    protein_name: str
    mutation_resolution: dict[str, Any]


class WorkbookProteinRecord(NamedTuple):
    """Logical protein entry parsed from an XLSX workbook."""
    protein_name: str
    structure_id: str
    chain: str | None
    mutations: list[dict[str, Any]]


# ---------------------------------------------------------------------------
# CSV parsing
# ---------------------------------------------------------------------------

def _normalise_column_name(name: str) -> str:
    """Case/spacing-insensitive normalisation for CSV header matching."""
    return "".join(ch for ch in name.lower() if ch.isalnum())


def _resolve_columns(
    fieldnames: list[str],
    aliases: dict[str, list[str]],
) -> dict[str, str]:
    """Map logical column names to actual CSV header names via aliases."""
    normalised_to_actual = {
        _normalise_column_name(name): name for name in fieldnames
    }

    resolved: dict[str, str] = {}
    missing: list[str] = []
    for logical_name, candidates in aliases.items():
        actual_name = None
        for candidate in candidates:
            actual_name = normalised_to_actual.get(_normalise_column_name(candidate))
            if actual_name is not None:
                break
        if actual_name is None:
            missing.append(logical_name)
        else:
            resolved[logical_name] = actual_name
    if missing:
        raise ValueError(
            f"CSV is missing required columns for fields {missing}. "
            f"Found headers: {fieldnames}."
        )
    return resolved


def _resolve_required_columns(fieldnames: list[str]) -> dict[str, str]:
    """
    Map logical column names to actual CSV header names.

    Supports compact public inference tables and workbook-exported mutation
    tables.
    """
    aliases = {
        "pdb": ["pdb", "pdb_file", "structure", "structure_id"],
        "chain": ["chain", "chain_id"],
        "position": ["position", "resnum", "res_num", "residue", "residuenumber"],
        "wt_aa": ["wt_aa", "wt", "wildtype", "wildtypeaa"],
        "mut_aa": ["mut_aa", "mt", "mut", "mutant", "mutantaa"],
    }
    try:
        resolved = _resolve_columns(fieldnames, aliases)
    except ValueError as exc:
        raise ValueError(
            "Mutations CSV is missing required columns. "
            f"{exc} Supported schemas include "
            "['pdb', 'chain', 'position', 'wt_aa', 'mut_aa']."
        ) from exc
    return resolved


def _resolve_structure_columns(fieldnames: list[str]) -> dict[str, str]:
    """Resolve the required public PDB and chain columns."""
    return _resolve_columns(
        fieldnames,
        {
            "pdb": ["pdb", "pdb_file", "structure", "structure_id"],
            "chain": ["chain", "chain_id"],
        },
    )


def resolve_dataset_csv_path(csv_path: str | Path) -> Path:
    """
    Resolve a dataset table path (CSV or XLSX).

    Existing relative paths are interpreted relative to the package root.
    Older package-relative table prefixes remain accepted for compatibility.

    The function is named `..._csv_path` for backward compatibility; it works
    with any table format whose readers are defined below.
    """
    path = Path(csv_path).expanduser()
    if not path.is_absolute():
        package_relative = (WORK_DIR / path).resolve()
        if package_relative.exists():
            return package_relative
    return resolve_path_with_base(
        csv_path,
        base_dir=INFERENCE_TABLE_BASE_DIR,
        strip_prefixes=(
            (WORK_DIR.name,),
        ),
    )


def _is_xlsx(path: Path) -> bool:
    return path.suffix.lower() == ".xlsx"


def _table_delimiter(path: Path) -> str:
    with path.open(encoding="utf-8") as handle:
        header = handle.readline()
    return "\t" if "\t" in header else ","


def _dict_reader(handle: Any, path: Path) -> csv.DictReader:
    return csv.DictReader(handle, delimiter=_table_delimiter(path))


# ---------------------------------------------------------------------------
# XLSX mutation-table readers
# ---------------------------------------------------------------------------

_TABLE_SHEET_NAME = "refined_sorted_clean"
MUTATION_INDEXING_MODE = "native_reference_exact"


def load_protein_names_csv(
    csv_path: str | Path,
    *,
    table_sheet_name: str = _TABLE_SHEET_NAME,
) -> list[str]:
    """
    Load unique internal structure keys derived from PDB and chain columns.
    """
    resolved = resolve_dataset_csv_path(csv_path)
    if not resolved.is_file():
        raise FileNotFoundError(f"Protein table not found: {resolved}")

    if _is_xlsx(resolved):
        return sorted(_iter_xlsx_protein_names(resolved, sheet_name=table_sheet_name))

    proteins: set[str] = set()
    with open(resolved, encoding="utf-8") as fh:
        reader = _dict_reader(fh, resolved)
        if reader.fieldnames is None:
            raise ValueError(f"Empty CSV: {resolved}")
        columns = _resolve_structure_columns(reader.fieldnames)
        for row in reader:
            pdb = str(row.get(columns["pdb"], "") or "").strip()
            chain = str(row.get(columns["chain"], "") or "").strip()
            if pdb and chain:
                proteins.add(structure_key(pdb, chain))

    return sorted(proteins)


def load_protein_structure_map(
    table_path: str | Path,
    *,
    table_sheet_name: str = _TABLE_SHEET_NAME,
) -> dict[str, str]:
    """
    Return internal PDB/chain key -> PDB filename stem.
    """
    resolved = resolve_dataset_csv_path(table_path)
    if not resolved.is_file():
        raise FileNotFoundError(f"Protein table not found: {resolved}")

    if _is_xlsx(resolved):
        return {
            protein_name: record.structure_id
            for protein_name, record in load_workbook_records_xlsx(
                resolved,
                sheet_name=table_sheet_name,
            ).items()
        }

    output: dict[str, str] = {}
    with open(resolved, encoding="utf-8") as fh:
        reader = _dict_reader(fh, resolved)
        if reader.fieldnames is None:
            raise ValueError(f"Empty CSV: {resolved}")
        columns = _resolve_structure_columns(reader.fieldnames)
        for row in reader:
            pdb = str(row.get(columns["pdb"], "") or "").strip()
            chain = str(row.get(columns["chain"], "") or "").strip()
            if not pdb or not chain:
                continue
            output[structure_key(pdb, chain)] = pdb_stem(pdb)
    return output


def load_protein_chain_map(
    table_path: str | Path,
    *,
    table_sheet_name: str = _TABLE_SHEET_NAME,
) -> dict[str, str]:
    """Return logical protein id -> author chain ID when the table has one."""
    resolved = resolve_dataset_csv_path(table_path)
    if not resolved.is_file():
        raise FileNotFoundError(f"Protein table not found: {resolved}")

    if _is_xlsx(resolved):
        return {
            protein_name: record.chain
            for protein_name, record in load_workbook_records_xlsx(
                resolved,
                sheet_name=table_sheet_name,
            ).items()
            if record.chain
        }

    with open(resolved, encoding="utf-8") as fh:
        reader = _dict_reader(fh, resolved)
        if reader.fieldnames is None:
            raise ValueError(f"Empty CSV: {resolved}")
        columns = _resolve_structure_columns(reader.fieldnames)

        output: dict[str, str] = {}
        for row in reader:
            pdb = str(row.get(columns["pdb"], "") or "").strip()
            chain = str(row.get(columns["chain"], "") or "").strip()
            if pdb and chain:
                output.setdefault(structure_key(pdb, chain), chain)
        return output


def _resolve_optional_column(fieldnames: list[str], candidates: list[str]) -> str | None:
    normalised_to_actual = {
        _normalise_column_name(name): name for name in fieldnames
    }
    for candidate in candidates:
        actual = normalised_to_actual.get(_normalise_column_name(candidate))
        if actual is not None:
            return actual
    return None


def _resolve_header_index(header: tuple[str, ...], candidates: list[str]) -> int | None:
    normalised_to_index = {
        _normalise_column_name(name): i
        for i, name in enumerate(header)
        if name
    }
    for candidate in candidates:
        idx = normalised_to_index.get(_normalise_column_name(candidate))
        if idx is not None:
            return idx
    return None


def _iter_xlsx_protein_names(
    xlsx_path: Path,
    *,
    sheet_name: str = _TABLE_SHEET_NAME,
) -> set[str]:
    return set(load_workbook_records_xlsx(xlsx_path, sheet_name=sheet_name).keys())


def load_workbook_records_xlsx(
    xlsx_path: str | Path,
    *,
    sheet_name: str = _TABLE_SHEET_NAME,
) -> dict[str, WorkbookProteinRecord]:
    """
    Parse an inference workbook into logical protein records.

    The internal structure key is derived from the required `pdb` and `chain`
    columns.
    """
    resolved = resolve_dataset_csv_path(xlsx_path)
    if not resolved.is_file():
        raise FileNotFoundError(f"Workbook not found: {resolved}")
    if not _is_xlsx(resolved):
        raise ValueError(f"Expected .xlsx workbook, got {resolved}.")

    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read XLSX mutation workbooks. "
            "Install it with `pip install openpyxl`."
        ) from exc

    required_aliases = {
        "pdb": ["pdb"],
        "chain": ["chain"],
        "wt_aa": ["wt_aa", "W2_AA"],
        "mt_aa": ["mt_aa", "MT_AA"],
        "mut_pos_pdb": ["mut_pos_pdb", "MUT_POS_PDB"],
    }
    records: dict[str, dict[str, Any]] = {}
    wb = openpyxl.load_workbook(resolved, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Workbook {resolved} does not contain the required "
                f"'{sheet_name}' sheet. Found: {wb.sheetnames}."
            )
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError(f"Sheet '{sheet_name}' in {resolved} is empty.")

        header = tuple(str(h) if h is not None else "" for h in header_row)
        idx: dict[str, int] = {}
        missing: list[str] = []
        for logical_name, aliases in required_aliases.items():
            header_idx = _resolve_header_index(header, aliases)
            if header_idx is None:
                missing.append(logical_name)
            else:
                idx[logical_name] = header_idx
        if missing:
            raise ValueError(
                f"Workbook {resolved} missing required columns {missing}. Found: {header}."
            )
        optional_idx = {
            "wt_sequence": _resolve_header_index(header, ["wt_seq", "WT_SEQ", "wildtype_sequence"]),
            "mut_pos_seq": _resolve_header_index(header, ["mut_pos_seq", "MUT_POS_SEQ"]),
        }

        for row in rows_iter:
            if idx["pdb"] >= len(row) or idx["chain"] >= len(row):
                continue

            pdb_id = str(row[idx["pdb"]]).strip() if row[idx["pdb"]] is not None else ""
            chain_val = str(row[idx["chain"]]).strip() if row[idx["chain"]] is not None else None
            if not pdb_id or not chain_val:
                continue
            protein_name = structure_key(pdb_id, chain_val)
            structure_id = pdb_stem(pdb_id)

            record = records.get(protein_name)
            if record is None:
                record = {
                    "protein_name": protein_name,
                    "structure_id": structure_id,
                    "chain": chain_val,
                    "mutations": [],
                }
                records[protein_name] = record
            else:
                if record["structure_id"] != structure_id:
                    raise ValueError(
                        f"Workbook {resolved} maps logical protein '{protein_name}' to "
                        f"multiple structure ids: '{record['structure_id']}' and '{structure_id}'."
                    )
                if chain_val and record["chain"] and record["chain"] != chain_val:
                    raise ValueError(
                        f"Workbook {resolved} maps logical protein '{protein_name}' to "
                        f"multiple chains: '{record['chain']}' and '{chain_val}'."
                    )
                if chain_val and not record["chain"]:
                    record["chain"] = chain_val
            wt_aa = str(row[idx["wt_aa"]]).strip().upper() if row[idx["wt_aa"]] is not None else ""
            mut_aa = str(row[idx["mt_aa"]]).strip().upper() if row[idx["mt_aa"]] is not None else ""
            if wt_aa not in AA_TO_INDEX or mut_aa not in AA_TO_INDEX or wt_aa == mut_aa:
                continue

            try:
                position_raw_1based = int(row[idx["mut_pos_pdb"]])
                position = position_raw_1based - 1
            except (TypeError, ValueError):
                continue
            position_seq_1based = None
            if optional_idx["mut_pos_seq"] is not None and optional_idx["mut_pos_seq"] < len(row):
                try:
                    position_seq_1based = int(row[optional_idx["mut_pos_seq"]])
                except (TypeError, ValueError):
                    position_seq_1based = None
            prot_mutation_index = None

            record["mutations"].append(
                {
                    "position": position,
                    "position_raw_1based": position_raw_1based,
                    "position_source": "pdb_residue_number",
                    "position_pdb": position_raw_1based,
                    "position_seq": position_seq_1based,
                    "pdb_resseq": position_raw_1based,
                    "pdb_ins_code": "",
                    "wt_aa": wt_aa,
                    "mut_aa": mut_aa,
                    "prot_mutation_index": prot_mutation_index,
                    "structure_id": structure_id,
                    "protein_name": protein_name,
                    "chain": record["chain"],
                    "wt_sequence": (
                        str(row[optional_idx["wt_sequence"]]).strip().upper()
                        if optional_idx["wt_sequence"] is not None
                        and optional_idx["wt_sequence"] < len(row)
                        and row[optional_idx["wt_sequence"]] is not None
                        else None
                    ),
                }
            )
    finally:
        wb.close()

    output: dict[str, WorkbookProteinRecord] = {}
    for protein_name, record in records.items():
        mutations = sorted(
            record["mutations"],
            key=lambda m: (
                m["position"],
                m["prot_mutation_index"] if m["prot_mutation_index"] is not None else 10**12,
                m["mut_aa"],
            ),
        )
        output[protein_name] = WorkbookProteinRecord(
            protein_name=protein_name,
            structure_id=record["structure_id"],
            chain=record["chain"],
            mutations=mutations,
        )
    return output


def load_mutations_xlsx(
    xlsx_path: str | Path,
    *,
    sheet_name: str = _TABLE_SHEET_NAME,
) -> dict[str, list[dict[str, Any]]]:
    """
    Load an XLSX mutation workbook and group rows by the key derived from its
    PDB and chain columns.

    Returns
    -------
    dict mapping protein_name -> list of mutation dicts with keys
        position (0-indexed int), wt_aa, and mut_aa.
    """
    return {
        protein_name: record.mutations
        for protein_name, record in load_workbook_records_xlsx(
            xlsx_path,
            sheet_name=sheet_name,
        ).items()
    }

def load_mutations_csv(
    csv_path: str | Path,
    *,
    table_sheet_name: str = _TABLE_SHEET_NAME,
) -> dict[str, list[dict[str, Any]]]:
    """
    Load a mutations table and group rows by the derived PDB/chain key.

    Returns
    -------
    protein_mutations : dict mapping PDB/chain key -> list of mutation dicts.
        Each mutation dict has keys for position, wild-type amino acid, and
        mutant amino acid.
    """
    resolved = resolve_dataset_csv_path(csv_path)
    if not resolved.is_file():
        raise FileNotFoundError(f"Mutations table not found: {resolved}")
    if _is_xlsx(resolved):
        return load_mutations_xlsx(resolved, sheet_name=table_sheet_name)
    csv_path = resolved

    protein_mutations: dict[str, list[dict[str, Any]]] = defaultdict(list)

    with open(csv_path, encoding="utf-8") as fh:
        reader = _dict_reader(fh, csv_path)

        if reader.fieldnames is None:
            raise ValueError(f"Empty CSV: {csv_path}")
        column_map = _resolve_required_columns(reader.fieldnames)

        for row_num, row in enumerate(reader, start=2):
            pdb = str(row.get(column_map["pdb"], "") or "").strip()
            chain = str(row.get(column_map["chain"], "") or "").strip()
            if not pdb or not chain:
                continue
            protein_name = structure_key(pdb, chain)

            try:
                position = int(row[column_map["position"]]) - 1  # convert 1-indexed to 0-indexed
            except ValueError:
                raise ValueError(
                    f"Row {row_num}: invalid position '{row[column_map['position']]}'"
                )

            wt_aa = row[column_map["wt_aa"]].strip().upper()
            mut_aa = row[column_map["mut_aa"]].strip().upper()
            if wt_aa not in AA_TO_INDEX or mut_aa not in AA_TO_INDEX:
                continue  # skip non-standard amino acids

            if wt_aa == mut_aa:
                continue  # skip identity "mutations"

            protein_mutations[protein_name].append({
                "position": position,
                "position_raw_1based": position + 1,
                "position_source": "model_index",
                "wt_aa": wt_aa,
                "mut_aa": mut_aa,
                "structure_id": pdb_stem(pdb),
                "chain": chain,
            })

    return dict(protein_mutations)


# ---------------------------------------------------------------------------
# Requested-mutation mask construction
# ---------------------------------------------------------------------------

def normalise_residue_mapping(
    residue_mapping: list[dict[str, Any]] | None,
    wt_sequence: str,
) -> list[dict[str, Any]]:
    """Return one residue-mapping record per continuous model position."""
    L = len(wt_sequence)
    if not residue_mapping:
        return [
            {
                "model_index": idx + 1,
                "aa": aa,
                "chain_id": None,
                "observed": None,
                "observed_index": None,
                "pdb_resseq": None,
                "pdb_ins_code": "",
                "pdb_residue": None,
                "pdb_residue_id": None,
            }
            for idx, aa in enumerate(wt_sequence)
        ]

    records_by_index: dict[int, dict[str, Any]] = {}
    for raw in residue_mapping:
        try:
            model_index = int(raw.get("model_index"))
        except (TypeError, ValueError):
            continue
        if model_index < 1 or model_index > L:
            continue
        record = dict(raw)
        record["model_index"] = model_index
        record["aa"] = str(record.get("aa") or wt_sequence[model_index - 1]).upper()
        if record.get("pdb_resseq") in {"", None}:
            record["pdb_resseq"] = None
        else:
            try:
                record["pdb_resseq"] = int(record["pdb_resseq"])
            except (TypeError, ValueError):
                record["pdb_resseq"] = None
        record["pdb_ins_code"] = str(record.get("pdb_ins_code") or "")
        records_by_index[model_index] = record

    output: list[dict[str, Any]] = []
    for idx, aa in enumerate(wt_sequence, start=1):
        output.append(
            records_by_index.get(
                idx,
                {
                    "model_index": idx,
                    "aa": aa,
                    "chain_id": None,
                    "observed": None,
                    "observed_index": None,
                    "pdb_resseq": None,
                    "pdb_ins_code": "",
                    "pdb_residue": None,
                    "pdb_residue_id": None,
                },
            )
        )
    return output


def _build_pdb_residue_lookup(
    residue_mapping: list[dict[str, Any]],
) -> dict[tuple[int, str], list[dict[str, Any]]]:
    lookup: dict[tuple[int, str], list[dict[str, Any]]] = defaultdict(list)
    for record in residue_mapping:
        resseq = record.get("pdb_resseq")
        if resseq is None:
            continue
        ins_code = str(record.get("pdb_ins_code") or "")
        lookup[(int(resseq), ins_code)].append(record)
    return dict(lookup)


def _global_align_reference_to_model(
    reference_sequence: str,
    model_sequence: str,
) -> list[int | None]:
    """Map 0-based reference-sequence positions to 0-based model-sequence positions."""
    if reference_sequence == model_sequence:
        return list(range(len(reference_sequence)))

    n = len(reference_sequence)
    m = len(model_sequence)
    gap = -2
    dp = [[0] * (m + 1) for _ in range(n + 1)]
    trace = [[""] * (m + 1) for _ in range(n + 1)]
    for i in range(1, n + 1):
        dp[i][0] = dp[i - 1][0] + gap
        trace[i][0] = "up"
    for j in range(1, m + 1):
        dp[0][j] = dp[0][j - 1] + gap
        trace[0][j] = "left"

    for i in range(1, n + 1):
        ref_aa = reference_sequence[i - 1]
        for j in range(1, m + 1):
            model_aa = model_sequence[j - 1]
            diag = dp[i - 1][j - 1] + (3 if ref_aa == model_aa else -3)
            up = dp[i - 1][j] + gap
            left = dp[i][j - 1] + gap
            best = max(diag, up, left)
            dp[i][j] = best
            trace[i][j] = "diag" if best == diag else ("up" if best == up else "left")

    ref_to_model: list[int | None] = [None] * n
    i, j = n, m
    while i > 0 or j > 0:
        direction = trace[i][j]
        if direction == "diag":
            ref_to_model[i - 1] = j - 1
            i -= 1
            j -= 1
        elif direction == "up" or j == 0:
            i -= 1
        else:
            j -= 1
    return ref_to_model


def _resolve_mutation_model_position_reference_parser(
    mutation: dict[str, Any],
    wt_sequence: str,
    parser_context: dict[str, Any] | None,
) -> tuple[int | None, str, str | None]:
    """Resolve mutation position with the native reference parser."""
    if parser_context is None:
        return (
            None,
            "native_reference_parser_missing",
            "native_reference_parser_missing",
        )

    idx, method = resolve_reference_mutation_index(
        mutation,
        parser_context=parser_context,
    )
    if idx is None:
        return None, method, method
    if idx < 0 or idx >= len(wt_sequence):
        return (
            None,
            f"native_reference_out_of_saprot_range:{method}",
            f"native_reference_out_of_saprot_range:{method}:idx_{idx}_len_{len(wt_sequence)}",
        )
    wt_aa = mutation["wt_aa"]
    if wt_sequence[idx] != wt_aa:
        return (
            None,
            f"native_reference_to_saprot_wt_mismatch:{method}",
            f"native_reference_to_saprot_wt_mismatch:{method}:saw_{wt_sequence[idx]}_expected_{wt_aa}",
        )
    return idx, f"native_reference:{method}", None


def _resolve_mutation_model_position(
    mutation: dict[str, Any],
    wt_sequence: str,
    residue_mapping: list[dict[str, Any]],
    pdb_lookup: dict[tuple[int, str], list[dict[str, Any]]],
    alignment_cache: dict[str, list[int | None]],
    parser_context: dict[str, Any] | None = None,
) -> tuple[int | None, str, str | None]:
    if MUTATION_INDEXING_MODE == "native_reference_exact":
        if "position_pdb" in mutation or mutation.get("position_source") in {"pdb_residue_number", "pdb_author_residue_number"}:
            return _resolve_mutation_model_position_reference_parser(
                mutation,
                wt_sequence,
                parser_context,
            )
        idx = int(mutation.get("position", -1))
        wt_aa_direct = mutation["wt_aa"]
        if 0 <= idx < len(wt_sequence) and wt_sequence[idx] == wt_aa_direct:
            return idx, "model_index:csv_direct", None
        return (
            None,
            "model_index:csv_direct_wt_mismatch_or_out_of_range",
            f"model_index:csv_direct_wt_mismatch_or_out_of_range:idx_{idx}_expected_{wt_aa_direct}",
        )

    wt_aa = mutation["wt_aa"]
    raw_position = int(mutation.get("position_raw_1based", mutation["position"] + 1))
    source = mutation.get("position_source", "model_index")
    ins_code = str(mutation.get("pdb_ins_code") or "")
    attempts: list[str] = []

    def check_model_index(model_index_1based: int, method: str) -> tuple[int | None, str | None]:
        idx = model_index_1based - 1
        if idx < 0 or idx >= len(wt_sequence):
            return None, f"{method}:out_of_range"
        if wt_sequence[idx] != wt_aa:
            return None, f"{method}:wt_mismatch_saw_{wt_sequence[idx]}_expected_{wt_aa}"
        return idx, None

    def try_pdb_number() -> tuple[int | None, str | None]:
        candidates = pdb_lookup.get((raw_position, ins_code), [])
        if not candidates and not ins_code:
            candidates = [
                record
                for (resseq, _candidate_ins), records in pdb_lookup.items()
                if resseq == raw_position
                for record in records
            ]
        if not candidates:
            return None, "pdb_residue_number:not_found"

        matching: list[int] = []
        mismatches: list[str] = []
        for record in candidates:
            idx = int(record["model_index"]) - 1
            if idx < 0 or idx >= len(wt_sequence):
                continue
            if wt_sequence[idx] == wt_aa:
                matching.append(idx)
            else:
                mismatches.append(f"{record.get('pdb_residue_id')}:{wt_sequence[idx]}")
        if len(matching) == 1:
            return matching[0], None
        if len(matching) > 1:
            return None, "pdb_residue_number:ambiguous_wt_match"
        return None, "pdb_residue_number:wt_mismatch_" + ",".join(mismatches[:5])

    if source == "pdb_residue_number" and pdb_lookup:
        idx, reason = try_pdb_number()
        if idx is not None:
            return idx, "pdb_residue_number", None
        attempts.append(reason or "pdb_residue_number:unresolved")

    idx, reason = check_model_index(raw_position, "model_index")
    if idx is not None:
        return idx, "model_index", None
    attempts.append(reason or "model_index:unresolved")

    if source != "pdb_residue_number" and pdb_lookup:
        idx, reason = try_pdb_number()
        if idx is not None:
            return idx, "pdb_residue_number", None
        attempts.append(reason or "pdb_residue_number:unresolved")

    workbook_sequence = mutation.get("wt_sequence")
    if workbook_sequence:
        workbook_sequence = str(workbook_sequence).strip().upper()
        if 1 <= raw_position <= len(workbook_sequence):
            if workbook_sequence[raw_position - 1] == wt_aa:
                ref_to_model = alignment_cache.get(workbook_sequence)
                if ref_to_model is None:
                    ref_to_model = _global_align_reference_to_model(
                        workbook_sequence,
                        wt_sequence,
                    )
                    alignment_cache[workbook_sequence] = ref_to_model
                mapped_idx = ref_to_model[raw_position - 1]
                if mapped_idx is not None and wt_sequence[mapped_idx] == wt_aa:
                    return mapped_idx, "workbook_sequence_alignment", None
                attempts.append("workbook_sequence_alignment:wt_mismatch_or_gap")
            else:
                attempts.append(
                    "workbook_sequence_alignment:reference_wt_mismatch_"
                    f"saw_{workbook_sequence[raw_position - 1]}_expected_{wt_aa}"
                )

    return None, "unresolved", ";".join(attempts)


def build_mutation_mask(
    mutations: list[dict[str, Any]],
    wt_sequence: str,
    residue_mapping: list[dict[str, Any]] | None = None,
    *,
    parser_context: dict[str, Any] | None = None,
    return_resolution: bool = False,
) -> torch.Tensor | tuple[torch.Tensor, dict[str, Any]]:
    """
    Build a requested-mutation mask M (L x 20) from a list of substitutions.

    M(i, a) = 1 when position i -> amino acid a was requested.
    M(i, a_wt) = 0 for identity substitutions.

    This explicitly supports workbook structures where one protein
    has many mutation rows:
      - some sites have no mutations at all -> entire row remains masked out
      - some sites are mutated to only a subset of amino acids -> only those
        observed mutant channels are unmasked
      - duplicate requests for the same (site, mutant) map to one entry

    Parameters
    ----------
    mutations : list of dicts with keys: position, wt_aa, and mut_aa.
    wt_sequence : wild-type amino acid sequence.

    Returns
    -------
    mask   : (L, 20) binary mask.
    """
    L = len(wt_sequence)
    residue_mapping = normalise_residue_mapping(residue_mapping, wt_sequence)
    pdb_lookup = _build_pdb_residue_lookup(residue_mapping)
    alignment_cache: dict[str, list[int | None]] = {}

    mask = torch.zeros(L, NUM_AMINO_ACIDS)
    resolution_records: list[dict[str, Any]] = []
    resolved_by: Counter[str] = Counter()
    skipped_by_reason: Counter[str] = Counter()
    for mut in mutations:
        pos, method, reason = _resolve_mutation_model_position(
            mut,
            wt_sequence,
            residue_mapping,
            pdb_lookup,
            alignment_cache,
            parser_context=parser_context,
        )
        if pos is None:
            skipped_by_reason[reason or method] += 1
            if return_resolution:
                resolution_records.append(
                    {
                        "protein_name": mut.get("protein_name"),
                        "raw_position_1based": mut.get(
                            "position_raw_1based",
                            mut.get("position", -1) + 1,
                        ),
                        "position_source": mut.get("position_source"),
                        "wt_aa": mut.get("wt_aa"),
                        "mut_aa": mut.get("mut_aa"),
                        "resolved": False,
                        "resolution_method": method,
                        "skip_reason": reason,
                    }
                )
            continue
        mut_idx = AA_TO_INDEX.get(mut["mut_aa"])
        wt_idx = AA_TO_INDEX.get(mut["wt_aa"])
        if mut_idx is None or wt_idx is None:
            continue

        # Never set mask for WT channel
        if mut_idx == wt_idx:
            continue

        mask[pos, mut_idx] = 1.0
        resolved_by[method] += 1
        if return_resolution:
            residue_record = residue_mapping[pos]
            resolution_records.append(
                {
                    "protein_name": mut.get("protein_name"),
                    "raw_position_1based": mut.get(
                        "position_raw_1based",
                        mut.get("position", -1) + 1,
                    ),
                    "position_source": mut.get("position_source"),
                    "model_index": pos + 1,
                    "pdb_residue_id": residue_record.get("pdb_residue_id"),
                    "pdb_resseq": residue_record.get("pdb_resseq"),
                    "pdb_ins_code": residue_record.get("pdb_ins_code"),
                    "wt_aa": mut.get("wt_aa"),
                    "mut_aa": mut.get("mut_aa"),
                    "prot_mutation_index": mut.get("prot_mutation_index"),
                    "structure_id": mut.get("structure_id"),
                    "chain": mut.get("chain"),
                    "resolved": True,
                    "resolution_method": method,
                    "skip_reason": None,
                }
            )

    if return_resolution:
        return mask, {
            "mutation_indexing_mode": MUTATION_INDEXING_MODE,
            "structure_parser": (
                {
                    "parser": parser_context.get("parser"),
                    "chain_used": parser_context.get("chain_used"),
                    "pdb_path": parser_context.get("pdb_path"),
                    "parser_error": parser_context.get("parser_error"),
                    "chain_candidates": parser_context.get("chain_candidates"),
                    "sequence_length": len(str(parser_context.get("seq", ""))),
                    "resn_list_length": len(parser_context.get("resn_list", [])),
                }
                if parser_context is not None
                else None
            ),
            "total_rows": len(mutations),
            "resolved_rows": sum(resolved_by.values()),
            "masked_entries": int(mask.sum().item()),
            "skipped_rows": sum(skipped_by_reason.values()),
            "resolved_by": dict(resolved_by),
            "skipped_by_reason": dict(skipped_by_reason),
            "records": resolution_records,
        }
    return mask


def _resolve_emb_path(protein_dir: Path) -> Path | None:
    """Return the canonical per-residue embedding path (new layout first)."""
    candidates = [
        protein_dir / "outputs" / "embeddings" / "per_residue_embeddings.pt",
        protein_dir / "embeddings" / "per_residue_embeddings.pt",
    ]
    for c in candidates:
        if c.is_file():
            return c
    return None


def _resolve_ca_coordinates_path(protein_dir: Path) -> Path | None:
    """Return the optional cached C-alpha coordinate tensor path."""
    candidates = [
        protein_dir / "inputs" / "numeric" / "ca_coordinates.pt",
        protein_dir / "outputs" / "structure" / "ca_coordinates.pt",
        protein_dir / "ca_coordinates.pt",
    ]
    for path in candidates:
        if path.is_file():
            return path
    return None


def _load_ca_coordinates(
    *,
    protein_dir: Path,
    length: int,
    residue_mapping: list[dict[str, Any]],
) -> torch.Tensor:
    """Load or reconstruct C-alpha coordinates aligned to model residues."""
    path = _resolve_ca_coordinates_path(protein_dir)
    if path is not None:
        coords = torch.load(path, map_location="cpu", weights_only=True).float()
        if coords.ndim != 2 or tuple(coords.shape) != (length, 3):
            raise ValueError(
                f"C-alpha coordinate shape {tuple(coords.shape)} does not match "
                f"expected ({length}, 3) for {protein_dir.name}."
            )
        return coords

    coords = torch.full((length, 3), float("nan"), dtype=torch.float32)
    for record in residue_mapping:
        try:
            model_index = int(record.get("model_index", 0) or 0)
        except (TypeError, ValueError):
            continue
        if model_index < 1 or model_index > length:
            continue
        if not record.get("has_ca"):
            continue
        try:
            coords[model_index - 1] = torch.tensor(
                [
                    float(record["ca_x"]),
                    float(record["ca_y"]),
                    float(record["ca_z"]),
                ],
                dtype=torch.float32,
            )
        except (KeyError, TypeError, ValueError):
            continue
    return coords


def _resolve_cached_pdb_path(metadata: dict[str, Any]) -> Path | None:
    raw = metadata.get("pdb_path")
    if not raw:
        return None
    return resolve_path_with_base(
        raw,
        base_dir=WORK_DIR,
        strip_prefixes=((WORK_DIR.name,),),
    )


def _build_sample_parser_context(sample: dict[str, Any]) -> dict[str, Any] | None:
    """Build native parser context for one cached SaProt sample."""
    if MUTATION_INDEXING_MODE != "native_reference_exact":
        return None
    pdb_path = sample.get("pdb_path")
    protein_name = str(sample.get("protein_name"))
    structure_id = str(sample.get("structure_id") or protein_name)
    if pdb_path is None:
        return {
            "parser": "SAAFEC_STAIR_native_pdb_chain_parser",
            "parser_error": "missing_cached_pdb_path",
            "seq": "",
            "resn_list": [],
        }
    pdb_path = Path(pdb_path)
    chain_candidates = unique_chain_candidates(
        sample.get("chain"),
        infer_chain_from_stem(protein_name),
        infer_chain_from_stem(structure_id),
        safe_first_atom_chain(pdb_path),
        "A",
    )
    try:
        return build_structure_parser_context(
            pdb_path=pdb_path,
            protein_name=protein_name,
            chain_candidates=chain_candidates,
        )
    except Exception as exc:  # noqa: BLE001 - keep sample loading tolerant.
        return {
            "parser": "SAAFEC_STAIR_native_pdb_chain_parser",
            "parser_error": repr(exc),
            "pdb_path": str(pdb_path),
            "chain_candidates": chain_candidates,
            "seq": "",
            "resn_list": [],
        }


class MutationDataset(Dataset):
    """Load cached SaProt embeddings for requested substitutions.

    Parameters
    ----------
    mutations_table : str or Path
        Path to the mutations table (CSV or XLSX). Existing relative paths are
        resolved under the package root. Supported schemas:
          * CSV: ``pdb, chain, position, wt_aa, mut_aa``.
          * XLSX: sheet ``refined_sorted_clean`` with columns pdb, chain,
            wt_aa, mt_aa, and mut_pos_pdb.
    embeddings_dir : str or Path
        Path to the by_protein/ directory containing cached embeddings. Each
        protein subdirectory must have outputs/embeddings/per_residue_embeddings.pt
        (or the legacy embeddings/per_residue_embeddings.pt) and metadata.json.
    mutations_csv : str or Path, optional
        Back-compat alias for ``mutations_table``.
    """

    def __init__(
        self,
        mutations_table: str | Path | None = None,
        embeddings_dir: str | Path | None = None,
        table_sheet_name: str = _TABLE_SHEET_NAME,
        *,
        mutations_csv: str | Path | None = None,
    ) -> None:
        if mutations_table is None:
            mutations_table = mutations_csv
        if mutations_table is None:
            raise TypeError("MutationDataset requires `mutations_table` (or the legacy `mutations_csv`).")
        if embeddings_dir is None:
            raise TypeError("MutationDataset requires `embeddings_dir`.")

        self.embeddings_dir = resolve_output_path(embeddings_dir)
        all_mutations = load_mutations_csv(
            mutations_table,
            table_sheet_name=table_sheet_name,
        )

        self.samples: list[dict[str, Any]] = []
        skipped_no_emb = 0

        for protein_name, mutations in all_mutations.items():
            protein_dir = self.embeddings_dir / protein_name
            emb_path = _resolve_emb_path(protein_dir)
            meta_path = protein_dir / "metadata.json"

            if emb_path is None or not meta_path.is_file():
                skipped_no_emb += 1
                continue

            with open(meta_path, encoding="utf-8") as f:
                meta = json.load(f)
            wt_sequence = meta.get("sequence")
            if wt_sequence is None:
                skipped_no_emb += 1
                continue
            residue_mapping = normalise_residue_mapping(
                meta.get("residue_mapping"),
                wt_sequence,
            )
            structure_id = str(meta.get("structure_id") or protein_name)
            chain = (
                mutations[0].get("chain")
                if mutations and mutations[0].get("chain")
                else meta.get("chain_id")
            )
            pdb_path = _resolve_cached_pdb_path(meta)

            self.samples.append({
                "protein_name": protein_name,
                "structure_id": structure_id,
                "chain": chain,
                "mutations": mutations,
                "wt_sequence": wt_sequence,
                "residue_mapping": residue_mapping,
                "pdb_path": pdb_path,
                "emb_path": emb_path,
                "protein_dir": protein_dir,
            })

        self.samples.sort(key=lambda s: s["protein_name"])

        n_mutations = sum(len(s["mutations"]) for s in self.samples)
        print(
            "MutationDataset: "
            f"{len(self.samples)} proteins, {n_mutations} mutations "
            f"(skipped: {skipped_no_emb} no embeddings)"
        )

    def __len__(self) -> int:
        return len(self.samples)

    def __getitem__(self, idx: int) -> ProteinSample:
        sample = self.samples[idx]

        embedding = torch.load(
            sample["emb_path"], map_location="cpu", weights_only=True
        )

        wt_sequence = sample["wt_sequence"]
        L = len(wt_sequence)
        ca_coordinates = _load_ca_coordinates(
            protein_dir=sample["protein_dir"],
            length=L,
            residue_mapping=sample.get("residue_mapping", []),
        )

        if embedding.ndim != 2 or embedding.shape[0] != L:
            raise ValueError(
                f"Embedding shape {tuple(embedding.shape)} does not match "
                f"sequence length {L} for {sample['protein_name']}"
            )

        mutation_mask, mutation_resolution = build_mutation_mask(
            sample["mutations"],
            wt_sequence,
            residue_mapping=sample.get("residue_mapping"),
            parser_context=_build_sample_parser_context(sample),
            return_resolution=True,
        )

        return ProteinSample(
            embedding=embedding,
            ca_coordinates=ca_coordinates,
            mutation_mask=mutation_mask,
            wt_sequence=wt_sequence,
            protein_name=sample["protein_name"],
            mutation_resolution=mutation_resolution,
        )


# ---------------------------------------------------------------------------
# Collate function for DataLoader
# ---------------------------------------------------------------------------

def mutation_collate_fn(
    batch: list[ProteinSample],
) -> dict[str, Any]:
    """
    Custom collate for variable-length proteins.

    Pads embeddings and mutation masks to the max length in the batch.
    Returns a dict with padded tensors and metadata.

    Padding uses zeros for embeddings and mutation masks.
    """
    max_len = max(s.embedding.shape[0] for s in batch)
    d_model = batch[0].embedding.shape[1]
    B = len(batch)

    embeddings = torch.zeros(B, max_len, d_model)
    ca_coordinates = torch.full((B, max_len, 3), float("nan"))
    mutation_masks = torch.zeros(B, max_len, NUM_AMINO_ACIDS)
    lengths = []
    sequences = []
    names = []
    mutation_resolutions = []

    for i, sample in enumerate(batch):
        L = sample.embedding.shape[0]
        embeddings[i, :L] = sample.embedding
        ca_coordinates[i, :L] = sample.ca_coordinates
        mutation_masks[i, :L] = sample.mutation_mask
        lengths.append(L)
        sequences.append(sample.wt_sequence)
        names.append(sample.protein_name)
        mutation_resolutions.append(sample.mutation_resolution)

    return {
        "embeddings": embeddings,    # (B, L_max, embedding_dim)
        "ca_coordinates": ca_coordinates,  # (B, L_max, 3)
        "mutation_masks": mutation_masks,  # (B, L_max, 20)
        "lengths": lengths,          # list of int
        "sequences": sequences,      # list of str
        "names": names,              # list of str
        "mutation_resolutions": mutation_resolutions,
    }
