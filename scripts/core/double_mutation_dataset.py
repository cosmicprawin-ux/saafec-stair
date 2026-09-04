#!/usr/bin/env python3
"""Dataset helpers for specified double-mutation inference workbooks.

A double row represents a requested pair of substitutions.
"""
from __future__ import annotations

import csv
import json
import math
import re
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any, NamedTuple

import torch
from torch.utils.data import Dataset

from core.proteinmpnn_logits_cache import ProteinMPNNLogitsCache
from core.amino_acids import AA_TO_INDEX, NUM_AMINO_ACIDS
from core.input_table import mutation_list_reader
from core.structure_identifiers import pdb_stem, structure_key


ROOT = Path(__file__).resolve().parents[2]
DOUBLE_SHEET = "refined_sorted"
COMPACT_MUTATION_RE = re.compile(r"^([A-Za-z])([0-9]+)([A-Za-z])$")


@dataclass(frozen=True)
class DoubleMutationRecord:
    identifier: str
    protein_name: str
    pdb: str
    chain: str
    wt_sequence: str
    mt_sequence: str
    positions: tuple[int, int]  # zero-based model-sequence positions
    positions_raw: tuple[int, int]  # one-based workbook positions
    wt_aa: tuple[str, str]
    mt_aa: tuple[str, str]
    prot_index: int | None
    prot_mutation_index: int | None


@dataclass(frozen=True)
class DoubleMutationProteinRecord:
    protein_name: str
    pdb: str
    chain: str
    wt_sequence: str
    mutations: tuple[DoubleMutationRecord, ...]


class DoubleMutationProteinSample(NamedTuple):
    protein_name: str
    pdb: str
    chain: str
    wt_sequence: str
    embeddings: torch.Tensor  # (L, D)
    single_ddg: torch.Tensor  # (L, 20)
    proteinmpnn_logits: torch.Tensor  # (L, 20)
    proteinmpnn_mask: torch.Tensor  # (L,)
    ca_coordinates: torch.Tensor  # (L, 3), NaN rows allowed
    positions: torch.Tensor  # (N, 2), zero-based
    wt_indices: torch.Tensor  # (N, 2)
    mt_indices: torch.Tensor  # (N, 2)
    identifiers: list[str]
    mutation_rows: list[DoubleMutationRecord]


def resolve_work_path(path: str | Path) -> Path:
    path = Path(path).expanduser()
    if path.is_absolute():
        return path
    if path.parts[:1] == ("database",):
        return ROOT / path
    return ROOT / path


def torch_load_compatible(path: Path, *, map_location: str | torch.device = "cpu", weights_only: bool = True) -> Any:
    try:
        return torch.load(path, map_location=map_location, weights_only=weights_only)
    except TypeError:
        return torch.load(path, map_location=map_location)


def _normalise_header(name: object) -> str:
    return "" if name is None else str(name).strip()


def _header_index(headers: tuple[object, ...]) -> dict[str, int]:
    return {_normalise_header(value): idx for idx, value in enumerate(headers)}


def _parse_pair(value: object, *, field: str, identifier: str) -> tuple[str, str]:
    parts = str(value).strip().split(";") if value is not None else []
    if len(parts) != 2 or not parts[0] or not parts[1]:
        raise ValueError(f"{identifier}: expected two semicolon-separated values in {field}, got {value!r}")
    return parts[0].strip(), parts[1].strip()


def _parse_int_pair(value: object, *, field: str, identifier: str) -> tuple[int, int]:
    left, right = _parse_pair(value, field=field, identifier=identifier)
    try:
        return int(left), int(right)
    except ValueError as exc:
        raise ValueError(f"{identifier}: invalid integer pair in {field}: {value!r}") from exc


def _optional_header(idx: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in idx:
            return idx[name]
    return None


def _clean_sequence(value: object) -> str:
    return "" if value is None else str(value).strip().upper()


def _normalise_column_name(name: object) -> str:
    return "".join(ch for ch in str(name or "").lower() if ch.isalnum())


def _resolve_column(fieldnames: list[str], candidates: list[str], *, required: bool = False, label: str = "") -> str | None:
    normalised_to_actual = {_normalise_column_name(name): name for name in fieldnames}
    for candidate in candidates:
        actual = normalised_to_actual.get(_normalise_column_name(candidate))
        if actual is not None:
            return actual
    if required:
        raise ValueError(f"Could not resolve {label or candidates[0]} column from headers: {fieldnames}")
    return None


def _parse_compact_mutation(value: object, *, field: str, identifier: str) -> tuple[str, int, str]:
    token = str(value or "").strip().upper()
    match = COMPACT_MUTATION_RE.match(token)
    if match is None:
        raise ValueError(f"{identifier}: expected compact mutation like K7Q in {field}, got {value!r}")
    wt_aa, position, mt_aa = match.groups()
    return wt_aa, int(position), mt_aa


def _split_double_mutation(value: object, *, identifier: str) -> tuple[str, str]:
    text = str(value or "").strip()
    parts = [part.strip() for part in re.split(r"[;,:+\s]+", text) if part.strip()]
    if len(parts) != 2:
        raise ValueError(f"{identifier}: expected two mutations in double_mutation, got {value!r}")
    return parts[0], parts[1]


def _format_pair(left: object, right: object) -> str:
    return f"{str(left).strip()};{str(right).strip()}"


def load_double_mutation_delimited(path: Path) -> dict[str, DoubleMutationProteinRecord]:
    records: dict[str, dict[str, Any]] = {}
    with path.open(encoding="utf-8") as handle:
        reader = mutation_list_reader(handle, path)
        if reader.fieldnames is None:
            raise ValueError(f"Empty mutation table: {path}")

        pdb_col = _resolve_column(
            reader.fieldnames,
            ["pdb", "pdb_file", "structure_id", "structure"],
            required=True,
            label="pdb",
        )
        chain_col = _resolve_column(
            reader.fieldnames,
            ["chain", "chain_id"],
            required=True,
            label="chain",
        )
        identifier_col = _resolve_column(reader.fieldnames, ["identifier", "mutation_id", "id"])
        mutation1_col = _resolve_column(reader.fieldnames, ["mut_1", "mutation_1", "mutation1", "first_mutation"])
        mutation2_col = _resolve_column(reader.fieldnames, ["mut_2", "mutation_2", "mutation2", "second_mutation"])
        double_mutation_col = _resolve_column(reader.fieldnames, ["double_mutation", "mutation_pair", "mutations"])
        wt_aa_col = _resolve_column(reader.fieldnames, ["wt_aa", "wt"])
        mt_aa_col = _resolve_column(reader.fieldnames, ["mt_aa", "mut_aa", "mt", "mutant"])
        mut_pos_col = _resolve_column(reader.fieldnames, ["mut_pos", "position_pair", "positions"])
        if (mutation1_col is None or mutation2_col is None) and double_mutation_col is None and (
            wt_aa_col is None or mt_aa_col is None or mut_pos_col is None
        ):
            raise ValueError(
                f"{path} needs either mut_1/mut_2, double_mutation, "
                "or wt_aa/mt_aa/mut_pos columns."
            )

        for row_num, row in enumerate(reader, start=2):
            pdb_value = str(row.get(pdb_col or "", "") or "").strip()
            chain = str(row.get(chain_col or "", "")).strip()
            if not pdb_value or not chain:
                continue
            pdb = pdb_stem(pdb_value)
            protein_name = structure_key(pdb, chain)
            identifier = str(row.get(identifier_col or "", "")).strip() if identifier_col is not None else ""

            if mutation1_col is not None and mutation2_col is not None:
                mutation1 = str(row.get(mutation1_col, "")).strip()
                mutation2 = str(row.get(mutation2_col, "")).strip()
            elif double_mutation_col is not None:
                mutation1, mutation2 = _split_double_mutation(row.get(double_mutation_col), identifier=identifier or f"row_{row_num}")
            else:
                wt_pair = _parse_pair(row.get(wt_aa_col), field="wt_aa", identifier=identifier or f"row_{row_num}")
                mt_pair = _parse_pair(row.get(mt_aa_col), field="mt_aa", identifier=identifier or f"row_{row_num}")
                raw_positions = _parse_int_pair(row.get(mut_pos_col), field="mut_pos", identifier=identifier or f"row_{row_num}")
                mutation1 = f"{wt_pair[0]}{raw_positions[0]}{mt_pair[0]}"
                mutation2 = f"{wt_pair[1]}{raw_positions[1]}{mt_pair[1]}"

            wt1, pos1, mt1 = _parse_compact_mutation(mutation1, field="mutation_1", identifier=identifier or f"row_{row_num}")
            wt2, pos2, mt2 = _parse_compact_mutation(mutation2, field="mutation_2", identifier=identifier or f"row_{row_num}")
            if identifier == "":
                identifier = f"{pdb}_{mutation1}:{mutation2}"

            if any(aa not in AA_TO_INDEX for aa in (wt1, wt2, mt1, mt2)):
                continue

            protein = records.get(protein_name)
            if protein is None:
                protein = {
                    "protein_name": protein_name,
                    "pdb": pdb,
                    "chain": chain,
                    "wt_sequence": "",
                    "mutations": [],
                }
                records[protein_name] = protein
            else:
                if protein["pdb"] != pdb:
                    raise ValueError(f"{path}: inconsistent pdb for {protein_name}")
                if protein["chain"] != chain:
                    raise ValueError(f"{path}: inconsistent chain for {protein_name}")

            protein["mutations"].append(
                DoubleMutationRecord(
                    identifier=identifier,
                    protein_name=protein_name,
                    pdb=pdb,
                    chain=chain,
                    wt_sequence="",
                    mt_sequence="",
                    positions=(pos1 - 1, pos2 - 1),
                    positions_raw=(pos1, pos2),
                    wt_aa=(wt1, wt2),
                    mt_aa=(mt1, mt2),
                    prot_index=None,
                    prot_mutation_index=None,
                )
            )

    output: dict[str, DoubleMutationProteinRecord] = {}
    for protein_name, record in records.items():
        mutations = tuple(
            sorted(
                record["mutations"],
                key=lambda item: (
                    item.positions[0],
                    item.positions[1],
                    item.mt_aa[0],
                    item.mt_aa[1],
                    item.identifier,
                ),
            )
        )
        output[protein_name] = DoubleMutationProteinRecord(
            protein_name=protein_name,
            pdb=record["pdb"],
            chain=record["chain"],
            wt_sequence="",
            mutations=mutations,
        )
    return output


def load_double_mutation_workbook(
    workbook_path: str | Path,
    *,
    sheet_name: str = DOUBLE_SHEET,
    validate_wt_positions: bool = True,
) -> dict[str, DoubleMutationProteinRecord]:
    """Parse a double-mutation workbook and group rows by PDB and chain."""
    path = resolve_work_path(workbook_path)
    if not path.is_file():
        raise FileNotFoundError(f"Double-mutation workbook not found: {path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return load_double_mutation_delimited(path)

    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise ImportError("openpyxl is required to read double-mutation workbooks.") from exc

    required = [
        "identifier",
        "pdb",
        "chain",
        "wt_aa",
        "mt_aa",
        "mut_pos",
    ]

    records: dict[str, dict[str, Any]] = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"{path} has no sheet {sheet_name!r}; found {wb.sheetnames}")
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = tuple(next(rows))
        except StopIteration:
            raise ValueError(f"{path}:{sheet_name} is empty.")
        idx = _header_index(headers)
        missing = [name for name in required if name not in idx]
        if missing:
            raise ValueError(f"{path}:{sheet_name} missing columns {missing}; found {headers}")
        wt_seq_idx = _optional_header(idx, "wt_seq", "wt_sequence", "sequence", "wildtype_sequence")
        mt_seq_idx = _optional_header(idx, "mt_seq", "mutant_sequence")

        for row in rows:
            identifier = str(row[idx["identifier"]]).strip() if row[idx["identifier"]] is not None else ""
            pdb_value = str(row[idx["pdb"]] or "").strip()
            chain = str(row[idx["chain"]] or "").strip()
            if not identifier or not pdb_value or not chain:
                continue
            pdb = pdb_stem(pdb_value)
            protein_name = structure_key(pdb, chain)
            wt_sequence = _clean_sequence(row[wt_seq_idx]) if wt_seq_idx is not None else ""
            mt_sequence = _clean_sequence(row[mt_seq_idx]) if mt_seq_idx is not None else ""
            wt_pair = tuple(aa.upper() for aa in _parse_pair(row[idx["wt_aa"]], field="wt_aa", identifier=identifier))
            mt_pair = tuple(aa.upper() for aa in _parse_pair(row[idx["mt_aa"]], field="mt_aa", identifier=identifier))
            raw_positions = _parse_int_pair(row[idx["mut_pos"]], field="mut_pos", identifier=identifier)
            positions = (raw_positions[0] - 1, raw_positions[1] - 1)
            if any(aa not in AA_TO_INDEX for aa in wt_pair + mt_pair):
                continue
            if wt_sequence and any(pos < 0 or pos >= len(wt_sequence) for pos in positions):
                raise ValueError(f"{identifier}: mutation positions {raw_positions} exceed sequence length {len(wt_sequence)}")
            if validate_wt_positions and wt_sequence:
                observed = (wt_sequence[positions[0]], wt_sequence[positions[1]])
                if observed != wt_pair:
                    raise ValueError(
                        f"{identifier}: wt_seq has {observed} at {raw_positions}, "
                        f"but wt_aa says {wt_pair}."
                    )

            protein = records.get(protein_name)
            if protein is None:
                protein = {
                    "protein_name": protein_name,
                    "pdb": pdb,
                    "chain": chain,
                    "wt_sequence": wt_sequence,
                    "mutations": [],
                }
                records[protein_name] = protein
            else:
                if protein["wt_sequence"] and wt_sequence and protein["wt_sequence"] != wt_sequence:
                    raise ValueError(f"{path}: inconsistent wt_seq for {protein_name}")
                if not protein["wt_sequence"] and wt_sequence:
                    protein["wt_sequence"] = wt_sequence
                if protein["pdb"] != pdb:
                    raise ValueError(f"{path}: inconsistent pdb for {protein_name}")
                if protein["chain"] != chain:
                    raise ValueError(f"{path}: inconsistent chain for {protein_name}")

            protein["mutations"].append(
                DoubleMutationRecord(
                    identifier=identifier,
                    protein_name=protein_name,
                    pdb=pdb,
                    chain=chain,
                    wt_sequence=wt_sequence,
                    mt_sequence=mt_sequence,
                    positions=positions,
                    positions_raw=raw_positions,
                    wt_aa=wt_pair,  # type: ignore[arg-type]
                    mt_aa=mt_pair,  # type: ignore[arg-type]
                    prot_index=None,
                    prot_mutation_index=None,
                )
            )
    finally:
        wb.close()

    output: dict[str, DoubleMutationProteinRecord] = {}
    for protein_name, record in records.items():
        mutations = tuple(
            sorted(
                record["mutations"],
                key=lambda item: (
                    item.positions[0],
                    item.positions[1],
                    item.mt_aa[0],
                    item.mt_aa[1],
                    item.identifier,
                ),
            )
        )
        output[protein_name] = DoubleMutationProteinRecord(
            protein_name=protein_name,
            pdb=record["pdb"],
            chain=record["chain"],
            wt_sequence=record["wt_sequence"],
            mutations=mutations,
        )
    return output


def load_pdb_derived_wt_sequence(protein_dir: Path, protein_name: str) -> str:
    metadata_path = protein_dir / "metadata.json"
    if not metadata_path.is_file():
        raise FileNotFoundError(f"{protein_name}: missing PDB-derived SaProt metadata: {metadata_path}")
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    sequence = _clean_sequence(metadata.get("sequence"))
    if not sequence:
        raise ValueError(f"{protein_name}: metadata.json does not contain a PDB-derived sequence")
    invalid = sorted({aa for aa in sequence if aa not in AA_TO_INDEX})
    if invalid:
        raise ValueError(f"{protein_name}: PDB-derived sequence contains unsupported residues: {''.join(invalid)}")
    return sequence


def with_pdb_derived_wt_sequence(
    record: DoubleMutationProteinRecord,
    wt_sequence: str,
    *,
    sequence_source: str,
) -> DoubleMutationProteinRecord:
    mutations: list[DoubleMutationRecord] = []
    for mutation in record.mutations:
        if any(pos < 0 or pos >= len(wt_sequence) for pos in mutation.positions):
            raise ValueError(
                f"{mutation.identifier}: mutation positions {mutation.positions_raw} exceed "
                f"PDB-derived sequence length {len(wt_sequence)} from {sequence_source}"
            )
        observed = (wt_sequence[mutation.positions[0]], wt_sequence[mutation.positions[1]])
        if observed != mutation.wt_aa:
            raise ValueError(
                f"{mutation.identifier}: PDB-derived sequence has {observed} at "
                f"{mutation.positions_raw}, but wt_aa says {mutation.wt_aa}."
            )
        mutations.append(replace(mutation, wt_sequence=wt_sequence))
    return replace(record, wt_sequence=wt_sequence, mutations=tuple(mutations))


def workbook_summary(workbook_path: str | Path, *, sheet_name: str = DOUBLE_SHEET) -> dict[str, Any]:
    records = load_double_mutation_workbook(workbook_path, sheet_name=sheet_name)
    row_count = sum(len(record.mutations) for record in records.values())
    return {
        "workbook": str(resolve_work_path(workbook_path)),
        "sheet": sheet_name,
        "proteins": len(records),
        "rows": row_count,
    }


def _candidate_protein_dirs(embeddings_dir: Path, record: DoubleMutationProteinRecord) -> list[Path]:
    stems = [
        record.protein_name,
        record.pdb,
        f"{record.pdb}_{record.chain}" if record.chain else "",
    ]
    seen: set[str] = set()
    output: list[Path] = []
    for stem in stems:
        if not stem or stem in seen:
            continue
        seen.add(stem)
        output.append(embeddings_dir / stem)
    return output


def _resolve_embedding_path(protein_dir: Path) -> Path | None:
    candidates = [
        protein_dir / "outputs" / "embeddings" / "per_residue_embeddings.pt",
        protein_dir / "embeddings" / "per_residue_embeddings.pt",
    ]
    for path in candidates:
        if path.is_file():
            return path
    return None


def _resolve_ca_coordinates_path(protein_dir: Path) -> Path | None:
    candidates = [
        protein_dir / "inputs" / "numeric" / "ca_coordinates.pt",
        protein_dir / "outputs" / "structure" / "ca_coordinates.pt",
        protein_dir / "ca_coordinates.pt",
    ]
    for path in candidates:
        if path.is_file():
            return path
    return None


def _load_ca_coordinates(protein_dir: Path, length: int) -> torch.Tensor:
    path = _resolve_ca_coordinates_path(protein_dir)
    if path is None:
        return torch.full((length, 3), float("nan"), dtype=torch.float32)
    coords = torch_load_compatible(path, map_location="cpu", weights_only=True).float()
    if coords.ndim != 2 or tuple(coords.shape) != (length, 3):
        raise ValueError(f"{path} has shape {tuple(coords.shape)}, expected ({length}, 3)")
    return coords


def _resolve_single_ddg_path(single_ddg_dir: Path, protein_name: str) -> Path | None:
    for suffix in (".pt", ".pth", ".csv"):
        path = single_ddg_dir / f"{protein_name}{suffix}"
        if path.is_file():
            return path
    return None


def _resolve_proteinmpnn_logits_path(proteinmpnn_cache_dir: Path, protein_name: str) -> Path | None:
    protein_dir = proteinmpnn_cache_dir / protein_name
    candidates = [
        protein_dir / "proteinmpnn_logits.pt",
        protein_dir / "outputs" / "proteinmpnn_logits.pt",
        protein_dir / "outputs" / "proteinmpnn" / "proteinmpnn_logits.pt",
        proteinmpnn_cache_dir / f"{protein_name}.pt",
    ]
    for path in candidates:
        if path.is_file():
            return path
    return None


def load_single_ddg_matrix(path: Path) -> torch.Tensor:
    """Load an ``(L, 20)`` single-mutant baseline matrix from PT/PTH or CSV."""
    if path.suffix.lower() in {".pt", ".pth"}:
        value = torch_load_compatible(path, map_location="cpu", weights_only=True)
        if isinstance(value, dict):
            for key in ("ddg", "single_ddg", "predictions", "matrix"):
                if key in value:
                    value = value[key]
                    break
        matrix = torch.as_tensor(value, dtype=torch.float32)
    elif path.suffix.lower() == ".csv":
        rows: list[list[float]] = []
        with path.open(encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames and all(aa in reader.fieldnames for aa in AA_TO_INDEX):
                for row in reader:
                    rows.append([float(row[aa]) for aa in AA_TO_INDEX])
            else:
                handle.seek(0)
                raw_reader = csv.reader(handle)
                for raw_row in raw_reader:
                    if not raw_row:
                        continue
                    rows.append([float(value) for value in raw_row[:NUM_AMINO_ACIDS]])
        matrix = torch.tensor(rows, dtype=torch.float32)
    else:
        raise ValueError(f"Unsupported single-DDG matrix format: {path}")

    if matrix.ndim != 2 or matrix.shape[1] != NUM_AMINO_ACIDS:
        raise ValueError(f"{path} must contain an (L, 20) matrix, got {tuple(matrix.shape)}")
    return matrix


class DoubleMutationDataset(Dataset):
    """One sample per protein, with all double-mutant rows for that protein."""

    def __init__(
        self,
        workbook_path: str | Path,
        *,
        embeddings_dir: str | Path,
        single_ddg_dir: str | Path,
        proteinmpnn_cache_dir: str | Path,
        sheet_name: str = DOUBLE_SHEET,
        require_all_inputs: bool = True,
    ) -> None:
        self.workbook_path = resolve_work_path(workbook_path)
        self.embeddings_dir = resolve_work_path(embeddings_dir)
        self.single_ddg_dir = resolve_work_path(single_ddg_dir)
        self.proteinmpnn_cache_dir = resolve_work_path(proteinmpnn_cache_dir)
        self.proteinmpnn_cache = ProteinMPNNLogitsCache(self.proteinmpnn_cache_dir)
        self.records = load_double_mutation_workbook(self.workbook_path, sheet_name=sheet_name)

        self.samples: list[dict[str, Any]] = []
        skipped_missing_embeddings = 0
        skipped_missing_single_ddg = 0
        skipped_missing_proteinmpnn = 0
        missing_required_inputs: list[str] = []
        for record in self.records.values():
            protein_dir = None
            emb_path = None
            for candidate_dir in _candidate_protein_dirs(self.embeddings_dir, record):
                candidate_emb = _resolve_embedding_path(candidate_dir)
                if candidate_emb is not None:
                    protein_dir = candidate_dir
                    emb_path = candidate_emb
                    break
            single_ddg_path = _resolve_single_ddg_path(self.single_ddg_dir, record.protein_name)
            proteinmpnn_logits_path = _resolve_proteinmpnn_logits_path(
                self.proteinmpnn_cache.root,
                record.protein_name,
            )
            if emb_path is None or protein_dir is None:
                skipped_missing_embeddings += 1
                missing_required_inputs.append(f"{record.protein_name}: missing embedding cache")
                if require_all_inputs:
                    continue
            if protein_dir is not None:
                pdb_sequence = load_pdb_derived_wt_sequence(protein_dir, record.protein_name)
                record = with_pdb_derived_wt_sequence(
                    record,
                    pdb_sequence,
                    sequence_source=str(protein_dir / "metadata.json"),
                )
                self.records[record.protein_name] = record
            if single_ddg_path is None:
                skipped_missing_single_ddg += 1
                missing_required_inputs.append(f"{record.protein_name}: missing single-DDG matrix")
                if require_all_inputs:
                    continue
            if proteinmpnn_logits_path is None:
                skipped_missing_proteinmpnn += 1
                missing_required_inputs.append(f"{record.protein_name}: missing ProteinMPNN logits")
                if require_all_inputs:
                    continue
            self.samples.append(
                {
                    "record": record,
                    "protein_dir": protein_dir,
                    "emb_path": emb_path,
                    "single_ddg_path": single_ddg_path,
                    "proteinmpnn_logits_path": proteinmpnn_logits_path,
                }
            )

        if require_all_inputs and missing_required_inputs:
            preview = "; ".join(missing_required_inputs[:10])
            suffix = "" if len(missing_required_inputs) <= 10 else f"; ... and {len(missing_required_inputs) - 10} more"
            raise FileNotFoundError(
                f"Missing required double-mutation inputs: {preview}{suffix}. "
                "Run the single-DDG export step with the same embeddings cache before inference."
            )

        self.samples.sort(key=lambda item: item["record"].protein_name)
        total_rows = sum(len(item["record"].mutations) for item in self.samples)
        print(
            f"DoubleMutationDataset: {len(self.samples)} proteins, {total_rows} rows "
            f"(skipped: {skipped_missing_embeddings} no embeddings, "
            f"{skipped_missing_single_ddg} no single-DDG matrix, "
            f"{skipped_missing_proteinmpnn} no ProteinMPNN logits)"
        )

    def __len__(self) -> int:
        return len(self.samples)

    def __getitem__(self, idx: int) -> DoubleMutationProteinSample:
        item = self.samples[idx]
        record: DoubleMutationProteinRecord = item["record"]
        if (
            item["emb_path"] is None
            or item["protein_dir"] is None
            or item["single_ddg_path"] is None
            or item["proteinmpnn_logits_path"] is None
        ):
            raise RuntimeError(f"Missing cached inputs for {record.protein_name}")

        embeddings = torch_load_compatible(item["emb_path"], map_location="cpu", weights_only=True).float()
        single_ddg = load_single_ddg_matrix(item["single_ddg_path"]).float()
        length = len(record.wt_sequence)
        if embeddings.ndim != 2 or embeddings.shape[0] != length:
            raise ValueError(
                f"{record.protein_name}: embeddings shape {tuple(embeddings.shape)} "
                f"does not match sequence length {length}"
            )
        if single_ddg.shape[0] != length:
            raise ValueError(
                f"{record.protein_name}: single-DDG shape {tuple(single_ddg.shape)} "
                f"does not match sequence length {length}"
            )
        proteinmpnn_logits, proteinmpnn_mask = self.proteinmpnn_cache.load(
            record.protein_name,
            expected_length=length,
            wt_sequence=record.wt_sequence,
        )
        ca_coordinates = _load_ca_coordinates(item["protein_dir"], length)

        positions = torch.tensor([mutation.positions for mutation in record.mutations], dtype=torch.long)
        wt_indices = torch.tensor(
            [[AA_TO_INDEX[mutation.wt_aa[0]], AA_TO_INDEX[mutation.wt_aa[1]]] for mutation in record.mutations],
            dtype=torch.long,
        )
        mt_indices = torch.tensor(
            [[AA_TO_INDEX[mutation.mt_aa[0]], AA_TO_INDEX[mutation.mt_aa[1]]] for mutation in record.mutations],
            dtype=torch.long,
        )
        return DoubleMutationProteinSample(
            protein_name=record.protein_name,
            pdb=record.pdb,
            chain=record.chain,
            wt_sequence=record.wt_sequence,
            embeddings=embeddings,
            single_ddg=single_ddg,
            proteinmpnn_logits=proteinmpnn_logits,
            proteinmpnn_mask=proteinmpnn_mask,
            ca_coordinates=ca_coordinates,
            positions=positions,
            wt_indices=wt_indices,
            mt_indices=mt_indices,
            identifiers=[mutation.identifier for mutation in record.mutations],
            mutation_rows=list(record.mutations),
        )


def double_mutation_collate_fn(batch: list[DoubleMutationProteinSample]) -> dict[str, Any]:
    if len(batch) != 1:
        raise ValueError("Double-mutation inference requires one protein per batch.")
    sample = batch[0]
    return {
        "protein_name": sample.protein_name,
        "pdb": sample.pdb,
        "chain": sample.chain,
        "wt_sequence": sample.wt_sequence,
        "embeddings": sample.embeddings,
        "single_ddg": sample.single_ddg,
        "proteinmpnn_logits": sample.proteinmpnn_logits,
        "proteinmpnn_mask": sample.proteinmpnn_mask,
        "ca_coordinates": sample.ca_coordinates,
        "positions": sample.positions,
        "wt_indices": sample.wt_indices,
        "mt_indices": sample.mt_indices,
        "identifiers": sample.identifiers,
        "mutation_rows": sample.mutation_rows,
    }
